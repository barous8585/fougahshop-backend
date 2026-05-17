"""
routes/onedrive.py — Synchronisation automatique OneDrive
==========================================================
Chaque nouvelle commande → nouvelle ligne dans FougahShop_Commandes.xlsx

Variables d'environnement requises sur Render :
  MS_TENANT_ID     = <votre_tenant_id>
  MS_CLIENT_ID     = <votre_client_id>
  MS_CLIENT_SECRET = <votre_client_secret>
  MS_USER_ID       = <votre_user_id>
  MS_FILE_ID       = <votre_file_id>
"""

import os
import io
import json
import tempfile
from datetime import datetime
from typing import Optional

import httpx

# ── Config depuis variables d'environnement ───────────────────
TENANT_ID     = os.environ.get("MS_TENANT_ID",     "")
CLIENT_ID     = os.environ.get("MS_CLIENT_ID",     "")
CLIENT_SECRET = os.environ.get("MS_CLIENT_SECRET", "")
USER_ID       = os.environ.get("MS_USER_ID",       "")
FILE_ID       = os.environ.get("MS_FILE_ID",       "")

TOKEN_URL     = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
GRAPH_BASE    = "https://graph.microsoft.com/v1.0"

# Cache token en mémoire
_token_cache: dict = {}


# ══════════════════════════════════════════════════════════════
# AUTHENTIFICATION
# ══════════════════════════════════════════════════════════════

async def get_access_token() -> Optional[str]:
    """Récupère un token d'accès Microsoft Graph (cache 55 min)."""
    global _token_cache
    now = datetime.utcnow().timestamp()

    # Token encore valide
    if _token_cache.get("token") and _token_cache.get("expires_at", 0) > now:
        return _token_cache["token"]

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(TOKEN_URL, data={
                "grant_type":    "client_credentials",
                "client_id":     CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "scope":         "https://graph.microsoft.com/.default",
            })
        if resp.status_code != 200:
            print(f"[OneDrive] Erreur token: {resp.text}")
            return None

        data = resp.json()
        _token_cache = {
            "token":      data["access_token"],
            "expires_at": now + data.get("expires_in", 3600) - 60,
        }
        return _token_cache["token"]

    except Exception as e:
        print(f"[OneDrive] Erreur auth: {e}")
        return None


# ══════════════════════════════════════════════════════════════
# TÉLÉCHARGEMENT ET UPLOAD DU FICHIER EXCEL
# ══════════════════════════════════════════════════════════════

async def download_excel(token: str) -> Optional[bytes]:
    """Télécharge le fichier Excel depuis OneDrive."""
    url = f"{GRAPH_BASE}/users/{USER_ID}/drive/items/{FILE_ID}/content"
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(url, headers={"Authorization": f"Bearer {token}"})
        if resp.status_code == 200:
            return resp.content
        print(f"[OneDrive] Erreur download: {resp.status_code} {resp.text[:200]}")
        return None
    except Exception as e:
        print(f"[OneDrive] Erreur download: {e}")
        return None


async def upload_excel(token: str, content: bytes) -> bool:
    """Upload le fichier Excel modifié vers OneDrive."""
    url = f"{GRAPH_BASE}/users/{USER_ID}/drive/items/{FILE_ID}/content"
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.put(
                url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                content=content,
            )
        if resp.status_code in (200, 201):
            print("[OneDrive] ✅ Fichier mis à jour")
            return True
        print(f"[OneDrive] Erreur upload: {resp.status_code} {resp.text[:200]}")
        return False
    except Exception as e:
        print(f"[OneDrive] Erreur upload: {e}")
        return False


# ══════════════════════════════════════════════════════════════
# CALCUL COMMISSION
# ══════════════════════════════════════════════════════════════

def _commission_fcfa(total_eu: float) -> int:
    if total_eu <= 50:   return 3500
    if total_eu <= 100:  return 5000
    if total_eu <= 200:  return 7000
    if total_eu <= 500:  return 12000
    return 20000


def _commission_locale(total_eu: float, monnaie: str, taux_gnf: float) -> int:
    comm_fcfa = _commission_fcfa(total_eu)
    if monnaie == "GNF":
        return round(comm_fcfa * taux_gnf / 656)
    return comm_fcfa


# ══════════════════════════════════════════════════════════════
# AJOUT D'UNE LIGNE DE COMMANDE
# ══════════════════════════════════════════════════════════════

async def ajouter_commande_excel(commande: dict) -> bool:
    """
    Ajoute une nouvelle ligne dans le fichier Excel OneDrive.

    commande = {
        ref, client_nom, client_tel, client_pays,
        total_euro, monnaie, articles (list ou JSON string),
        statut, created_at, promo_code, note_admin,
        taux_gnf (optionnel), frais_livraison_boutique (optionnel)
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[OneDrive] openpyxl non installé")
        return False

    token = await get_access_token()
    if not token:
        return False

    # Télécharger le fichier
    content = await download_excel(token)
    if not content:
        return False

    try:
        # Charger le workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb["📋 Commandes"]

        # Trouver la première ligne vide (à partir de la ligne 4)
        next_row = 4
        for row in range(4, 10000):
            if ws.cell(row=row, column=2).value is None:
                next_row = row
                break

        # Préparer les données
        ref        = commande.get("ref", "")
        nom        = commande.get("client_nom", "")
        tel        = commande.get("client_tel", "")
        pays       = commande.get("client_pays", "")
        total_eu   = float(commande.get("total_euro", 0) or 0)
        monnaie    = commande.get("monnaie", "GNF")
        statut     = commande.get("statut", "En attente")
        note       = commande.get("note_admin", "") or ""
        promo      = commande.get("promo_code", "") or ""
        created_at = commande.get("created_at")
        taux_gnf   = float(commande.get("taux_gnf", 0) or 9500)

        # Taux selon monnaie
        taux = taux_gnf if monnaie == "GNF" else 656

        # Articles — résumé texte
        arts = commande.get("articles", [])
        if isinstance(arts, str):
            try:
                arts = json.loads(arts)
            except Exception:
                arts = []
        articles_txt = " | ".join([
            f"{a.get('nom','?')} ×{a.get('qty',1)} ({a.get('prix_eu',0)}€)"
            for a in arts
        ]) if arts else ""

        # Frais livraison boutique total
        frais_b_total = sum(
            float(a.get("frais_livraison_boutique", 0) or 0) * int(a.get("qty", 1))
            for a in arts
        ) if arts else 0

        # Prix article seul (sans frais boutique)
        prix_eu_seul = max(0, total_eu - frais_b_total)

        # Commission
        comm_locale = _commission_locale(total_eu, monnaie, taux_gnf)

        # Montant converti (sans commission)
        montant_converti = round(total_eu * taux)

        # Total client
        total_client = montant_converti + comm_locale

        # Date
        if isinstance(created_at, str):
            try:
                date_val = datetime.fromisoformat(created_at.replace("Z",""))
            except Exception:
                date_val = datetime.now()
        elif isinstance(created_at, datetime):
            date_val = created_at
        else:
            date_val = datetime.now()

        # Statut traduit
        statut_map = {
            "en_attente_paiement": "En attente",
            "paye":     "Payé",
            "achete":   "Acheté",
            "expedie":  "Expédié",
            "arrive":   "Arrivé",
            "recupere": "Récupéré",
            "annulee":  "Annulé",
        }
        statut_fr = statut_map.get(statut, statut)

        # Écriture des cellules
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        VERT  = "1A8C5F"
        OR    = "B8860B"
        ROUGE = "C0392B"
        VERT_CLAIR  = "E8F5EE"
        ROUGE_CLAIR = "FDECEA"
        bg = "FFFFFF" if next_row % 2 == 0 else "F7F7F7"

        def s(b="E0E0E0"):
            side = Side(border_style="thin", color=b)
            return Border(left=side, right=side, top=side, bottom=side)

        def w(row, col, val, fmt=None, bold=False, color="2D2D2D",
              bg_=None, h="left"):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", bold=bold, size=10, color=color)
            c.fill = PatternFill("solid", fgColor=bg_ or bg)
            c.alignment = Alignment(horizontal=h, vertical="center")
            c.border = s()
            if fmt: c.number_format = fmt
            return c

        # Col 1 : Date
        c = ws.cell(row=next_row, column=1, value=date_val)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = s()
        c.number_format = "DD/MM/YYYY"

        # Col 2 : Référence
        w(next_row, 2, ref, bold=True, color=VERT, h="center")

        # Col 3 : Nom client
        w(next_row, 3, nom)

        # Col 4 : Téléphone
        w(next_row, 4, tel, h="center")

        # Col 5 : Pays
        w(next_row, 5, pays)

        # Col 6 : Articles
        c = ws.cell(row=next_row, column=6, value=articles_txt)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = s()

        # Col 7 : Prix Europe
        w(next_row, 7, round(prix_eu_seul, 2), fmt='#,##0.00 "€"',
          color="1A5276", h="right")

        # Col 8 : Livr. boutique
        w(next_row, 8, round(frais_b_total, 2) if frais_b_total else None,
          fmt='#,##0.00 "€"', color="1A5276", h="right")

        # Col 9 : Total Europe (formule)
        r = str(next_row)
        ws.cell(row=next_row, column=9,
                value=f'=IF(G{r}="","",G{r}+IF(H{r}="",0,H{r}))')
        ws.cell(row=next_row, column=9).number_format = '#,##0.00 "€"'
        ws.cell(row=next_row, column=9).font = Font(name="Arial", size=10)
        ws.cell(row=next_row, column=9).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=next_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=next_row, column=9).border = s()

        # Col 10 : Taux
        w(next_row, 10, taux, fmt='#,##0', h="right")

        # Col 11 : Monnaie
        w(next_row, 11, monnaie, h="center")

        # Col 12 : Montant converti
        w(next_row, 12, montant_converti, fmt='#,##0',
          bold=True, color=VERT, h="right")

        # Col 13 : Commission
        w(next_row, 13, comm_locale, fmt='#,##0',
          bold=True, color=OR, h="right")

        # Col 14 : Total client
        bg_total = "D5EDE0" if next_row % 2 == 0 else VERT_CLAIR
        c = ws.cell(row=next_row, column=14, value=total_client)
        c.font = Font(name="Arial", bold=True, size=11, color=VERT)
        c.fill = PatternFill("solid", fgColor=bg_total)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = Border(
            left=Side(border_style="medium", color=VERT),
            right=Side(border_style="medium", color=VERT),
            top=Side(border_style="medium", color=VERT),
            bottom=Side(border_style="medium", color=VERT),
        )
        c.number_format = '#,##0'

        # Col 15 : Frais de port (vide, rempli après pesée)
        w(next_row, 15, None, h="right")

        # Col 16 : Statut
        w(next_row, 16, statut_fr, bold=True, h="center")

        # Col 17 : Date paiement (vide)
        c = ws.cell(row=next_row, column=17)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = s()
        c.number_format = "DD/MM/YYYY"

        # Col 18 : Note
        note_clean = note.replace("[PRIVE]", "").strip(" |") if note else ""
        if promo:
            note_clean = (f"Promo: {promo} | " + note_clean).strip(" |")
        w(next_row, 18, note_clean, wrap=True) if False else (
            lambda: (
                setattr(ws.cell(row=next_row, column=18), 'value', note_clean),
                setattr(ws.cell(row=next_row, column=18), 'font', Font(name="Arial", size=9, italic=True)),
                setattr(ws.cell(row=next_row, column=18), 'fill', PatternFill("solid", fgColor=bg)),
                setattr(ws.cell(row=next_row, column=18), 'alignment',
                        Alignment(horizontal="left", vertical="center", wrap_text=True)),
                setattr(ws.cell(row=next_row, column=18), 'border', s()),
            )()
        )
        c18 = ws.cell(row=next_row, column=18, value=note_clean)
        c18.font = Font(name="Arial", size=9, italic=True)
        c18.fill = PatternFill("solid", fgColor=bg)
        c18.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c18.border = s()

        # Col 19 : Je gagne (= commission)
        bg_gain = "FAD7D3" if next_row % 2 == 0 else "FDECEA"
        c = ws.cell(row=next_row, column=19, value=comm_locale)
        c.font = Font(name="Arial", bold=True, size=11, color="C0392B")
        c.fill = PatternFill("solid", fgColor=bg_gain)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = Border(
            left=Side(border_style="medium", color="C0392B"),
            right=Side(border_style="medium", color="C0392B"),
            top=Side(border_style="medium", color="C0392B"),
            bottom=Side(border_style="medium", color="C0392B"),
        )
        c.number_format = '#,##0'

        # Sauvegarder en mémoire
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        content_new = buf.read()

        # Uploader sur OneDrive
        return await upload_excel(token, content_new)

    except Exception as e:
        print(f"[OneDrive] Erreur ajout ligne: {e}")
        import traceback
        traceback.print_exc()
        return False


# ══════════════════════════════════════════════════════════════
# MISE À JOUR DU STATUT
# ══════════════════════════════════════════════════════════════

async def mettre_a_jour_statut(ref: str, nouveau_statut: str,
                                frais_port: Optional[int] = None) -> bool:
    """Met à jour le statut d'une commande existante dans Excel."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    token = await get_access_token()
    if not token:
        return False

    content = await download_excel(token)
    if not content:
        return False

    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb["📋 Commandes"]

        statut_map = {
            "en_attente_paiement": "En attente",
            "paye":     "Payé",
            "achete":   "Acheté",
            "expedie":  "Expédié",
            "arrive":   "Arrivé",
            "recupere": "Récupéré",
            "annulee":  "Annulé",
        }
        statut_fr = statut_map.get(nouveau_statut, nouveau_statut)

        # Chercher la ligne avec cette référence (col 2)
        for row in range(4, 10000):
            cell_ref = ws.cell(row=row, column=2).value
            if cell_ref is None:
                break
            if str(cell_ref).strip().upper() == ref.strip().upper():
                # Mettre à jour statut (col 16)
                ws.cell(row=row, column=16).value = statut_fr

                # Date paiement si payé (col 17)
                if nouveau_statut == "paye":
                    ws.cell(row=row, column=17).value = datetime.now()
                    ws.cell(row=row, column=17).number_format = "DD/MM/YYYY"

                # Frais de port si fournis (col 15)
                if frais_port:
                    ws.cell(row=row, column=15).value = frais_port
                    ws.cell(row=row, column=15).number_format = '#,##0'

                break

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return await upload_excel(token, buf.read())

    except Exception as e:
        print(f"[OneDrive] Erreur update statut: {e}")
        return False
